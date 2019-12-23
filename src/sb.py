import tkinter as tk 
import tkinter.filedialog as Dialog
import os
import json 
import translateapi
import openpyxl
import re

json_root = "E:/st_kf/doc/language"
excel_root = "E:/sy_translation"
def set_json_dir(root, file_content, filetypes = None, title = "Open", initialdir = None):
    filename=Dialog.askopenfilename(master = root, filetypes=filetypes, title = title, initialdir = initialdir)
    file_content.set(filename)

def fix_escape_help(root, file_1, file_2):
    jsonfilename = file_1.get()
    excelfilename = file_2.get()
    if jsonfilename == "" or not os.path.isfile(jsonfilename) or excelfilename == "" or not os.path.isfile(excelfilename):
        # MsgBox.showerror(u"错误", u"请选择一个json文件")
        return
    else:
        with open(jsonfilename, "r", encoding='UTF-8') as f:
            jdata = json.load(f)
        lanlist = jdata['data']
        cnset0 = translateapi.create_cn_set(lanlist)
        cnset = {}
        for k, v in cnset0.items():
            k = translateapi.escape_string(k)
            cnset[k] = v

        eset = {}
        for k in cnset:
            e = k.replace('\\\\', '\\')
            eset[e] = translateapi.invert_escape_string(k)
        wb = openpyxl.load_workbook(excelfilename)
        ws = wb[wb.sheetnames[0]]
        for i in ws.iter_rows():
            src = i[0].value
            if src != None:
                # 后台提取出来的/带有\，被json.load进入后会被删除
                # 为了保持一致，尝试删除原文中的\
                src = src.replace(r'\/', r'/')
                if not (src in cnset):
                    if src in eset:
                        src = translateapi.escape_string(eset[src])
                        print(src, " is fixed")
                        i[0].value = src 
        wb.save(excelfilename)


def fix_escape(root):
    wnd = tk.Frame(root)
    title1_1 = tk.Label(master=wnd, text = "提取json源:")
    file1_1 = tk.StringVar(value = "")
    input1_1 = tk.Entry(master=wnd, textvariable=file1_1, width = 100)
    btn1_1 = tk.Button(master=wnd, text="打开", command=lambda : set_json_dir(root, file1_1, filetypes=[("Json File", "*.json")], title="文件源json", initialdir=json_root))
    title1_2 = tk.Label(master=wnd, text = "修复译文:")
    file1_2 = tk.StringVar(value = "")
    input1_2 = tk.Entry(master=wnd, textvariable=file1_2, width = 100)
    btn1_2 = tk.Button(master=wnd, text="打开", command=lambda : set_json_dir(root, file1_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=excel_root))
    pickbtn = tk.Button(master=wnd, text="提取", command=lambda : fix_escape_help(root, file1_1, file1_2))
    row = 0    
    title1_1.grid(row = row, column = 0)
    input1_1.grid(row = row, column = 1, columnspan = 10)
    btn1_1.grid(row = row, column = 11)
    row += 1
    title1_2.grid(row = row, column = 1, columnspan = 10, sticky = tk.W)
    row += 1
    input1_2.grid(row = row, column = 1, columnspan = 10)
    btn1_2.grid(row = row, column = 11)
    row += 1
    pickbtn.grid(row = row, column = 1, sticky = tk.W)
    wnd.pack()

def fix_talk_help(root, file_2):
    excelfilename = file_2.get()
    if excelfilename == "" or not os.path.isfile(excelfilename):
        # MsgBox.showerror(u"错误", u"请选择一个json文件")
        return
    else:
        wb = openpyxl.load_workbook(excelfilename)
        ws = wb[wb.sheetnames[0]]
        rep = re.compile(r'([0-9]+),')
        for i in ws.iter_rows():
            src = i[0].value

            if src != None:
                if re.search(r':[0-9]+$', src) != None:
                    src2 = re.sub(rep,lambda x: x.group(1) + r'||', src)
                    dest = i[1].value
                    dest2 = re.sub(rep,lambda x: x.group(1) + r'||', dest)
                    i[0].value = src2 
                    i[1].value = dest2
                    print(src)
                    print(src2, dest2)
        wb.save(excelfilename)

def fix_talk(root):
    wnd = tk.Frame(root)
    title1_2 = tk.Label(master=wnd, text = "修复译文:")
    file1_2 = tk.StringVar(value = "")
    input1_2 = tk.Entry(master=wnd, textvariable=file1_2, width = 100)
    btn1_2 = tk.Button(master=wnd, text="打开", command=lambda : set_json_dir(root, file1_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=excel_root))
    pickbtn = tk.Button(master=wnd, text="提取", command=lambda : fix_talk_help(root, file1_2))
    row = 0    
    title1_2.grid(row = row, column = 1, columnspan = 10, sticky = tk.W)
    row += 1
    input1_2.grid(row = row, column = 1, columnspan = 10)
    btn1_2.grid(row = row, column = 11)
    row += 1
    pickbtn.grid(row = row, column = 1, sticky = tk.W)
    wnd.pack()

def excel_to_dict(ws):
    s = {}
    for i in ws.iter_rows():
        src = i[0].value
        dest = i[1].value 
        s[src] = dest
    return s

def diff_language(root, file1, file2):
    excel1 = file1.get()
    excel2 = file2.get()
    wb = openpyxl.load_workbook(excel1, read_only=True)
    ws = wb[wb.sheetnames[0]]
    dict1 = excel_to_dict(ws)
    wb.close()
    wb = openpyxl.load_workbook(excel2, read_only=True)
    ws = wb[wb.sheetnames[0]]
    dict2 = excel_to_dict(ws)
    wb.close()
    newdict = {}
    for k, v in dict1.items():
        if not k in dict2 or dict2[k] != v:
            newdict[k] = v 
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = u"中文"
    ws.cell(1, 2).value = u"译文"
    row = 2
    for k, v in newdict.items():
        ws.cell(row, 1).value = k
        ws.cell(row, 2).value = v
        row += 1
    (_, file_name) = os.path.split(excel1)
    (short_name, _) = os.path.splitext(file_name)
    save_dir = Dialog.asksaveasfilename(master = root, filetypes=[("Excel File", "*.xlsx")], title = "请选择保存Excel文件的路径", initialdir = excel_root, initialfile = short_name)
    if save_dir == "":
        return
    (save_dir, file_name) = os.path.split(save_dir)
    (short_name, _) = os.path.splitext(file_name)
    save_full_name = os.path.join(save_dir, short_name + ".xlsx")
    wb.save(save_full_name)
    wb.close()

def diff_lan(root):
    wnd = tk.Frame(root)
    title1_2 = tk.Label(master=wnd, text = "新Excel:")
    file1_2 = tk.StringVar(value = "")
    input1_2 = tk.Entry(master=wnd, textvariable=file1_2, width = 100)
    btn1_2 = tk.Button(master=wnd, text="打开", command=lambda : set_json_dir(root, file1_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=excel_root))
    title2_2 = tk.Label(master=wnd, text = "旧Excel:")
    file2_2 = tk.StringVar(value = "")
    input2_2 = tk.Entry(master=wnd, textvariable=file2_2, width = 100)
    btn2_2 = tk.Button(master=wnd, text="打开", command=lambda : set_json_dir(root, file2_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=excel_root))
    pickbtn = tk.Button(master=wnd, text="提取", command=lambda : diff_language(root, file1_2, file2_2))
    row = 0    
    title1_2.grid(row = row, column = 1, columnspan = 10, sticky = tk.W)
    row += 1
    input1_2.grid(row = row, column = 1, columnspan = 10)
    btn1_2.grid(row = row, column = 11)
    row += 1
    title2_2.grid(row = row, column = 1, columnspan = 10, sticky = tk.W)
    row += 1
    input2_2.grid(row = row, column = 1, columnspan = 10)
    btn2_2.grid(row = row, column = 11)
    row += 1
    pickbtn.grid(row = row, column = 1, sticky = tk.W)
    wnd.pack()

def fix_format_help(root, file_2):
    excelfilename = file_2.get()
    if excelfilename == "" or not os.path.isfile(excelfilename):
        # MsgBox.showerror(u"错误", u"请选择一个json文件")
        return
    else:
        wb = openpyxl.load_workbook(excelfilename)
        ws = wb[wb.sheetnames[0]]
        rep = re.compile(r'^[\s]+|[\s]+$')
        for i in ws.iter_rows():
            for k in range(1,2):
                src = i[k].value
                if src != None and type(src) == str:
                    if re.search(rep, src) != None:
                        src2 = re.sub(rep, "", src)
                        i[k].value = src2
                        print(src2)

        wb.save(excelfilename)

def fix_format(root):
    wnd = tk.Frame(root)
    title1_2 = tk.Label(master=wnd, text = "修复译文:")
    file1_2 = tk.StringVar(value = "")
    input1_2 = tk.Entry(master=wnd, textvariable=file1_2, width = 100)
    btn1_2 = tk.Button(master=wnd, text="打开", command=lambda : set_json_dir(root, file1_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=excel_root))
    pickbtn = tk.Button(master=wnd, text="提取", command=lambda : fix_format_help(root, file1_2))
    row = 0    
    title1_2.grid(row = row, column = 1, columnspan = 10, sticky = tk.W)
    row += 1
    input1_2.grid(row = row, column = 1, columnspan = 10)
    btn1_2.grid(row = row, column = 11)
    row += 1
    pickbtn.grid(row = row, column = 1, sticky = tk.W)
    wnd.pack()

if __name__ == "__main__":
    root = tk.Tk()
    root.title(u"翻译工具")
    # fix_escape(root)
    # fix_talk(root)
    diff_lan(root)
    # fix_format(root)
    root.mainloop()

