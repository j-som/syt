import tkinter as tk 
import tkinter.filedialog as Dialog
import tkinter.messagebox as MsgBox
import sytlog
import os
import json 
import translateapi
import openpyxl
import re
import viewtool
import config
import ui

def excel_to_dict(ws):
    s = {}
    for i in ws.iter_rows():
        src = i[0].value
        dest = i[1].value 
        s[src] = dest
    return s

class SlimLan(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master=master)
        title1_2 = tk.Label(master=self, text = "Excel:")
        file1_2 = tk.StringVar(value = "")
        input1_2 = tk.Entry(master=self, textvariable=file1_2, width = 80)
        btn1_2 = tk.Button(master=self, text="打开", command=lambda : viewtool.set_json_dir(self.winfo_toplevel(), file1_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=config.excel_root()))
        title2_2 = tk.Label(master=self, text = "Json:")
        file2_2 = tk.StringVar(value = "")
        input2_2 = tk.Entry(master=self, textvariable=file2_2, width = 80)
        btn2_2 = tk.Button(master=self, text="打开", command=lambda : viewtool.set_json_dir(self.winfo_toplevel(), file2_2, filetypes=[("Json File", "*.json")], title="中文json", initialdir=config.json_root()))
        pickbtn = tk.Button(master=self, text="清理", command=lambda : self.slim_language(file1_2, file2_2))
        row = 0    
        title1_2.grid(row = row, column = 0, columnspan = 6, sticky = tk.W)
        row += 1
        input1_2.grid(row = row, column = 0, columnspan = 6)
        btn1_2.grid(row = row, column = 6, sticky = tk.EW, padx = 5)
        row += 1
        title2_2.grid(row = row, column = 0, columnspan = 6, sticky = tk.W)
        row += 1
        input2_2.grid(row = row, column = 0, columnspan = 6)
        btn2_2.grid(row = row, column = 6, sticky = tk.EW, padx = 5)
        row += 1
        pickbtn.grid(row = row, column = 0, sticky = tk.W)



    def slim_language(self, file1, file2):
        excel1 = file1.get()
        jsonfilename = file2.get()
        wb = openpyxl.load_workbook(excel1, read_only=True)
        ws = wb[wb.sheetnames[0]]
        dict1 = translateapi.excel_to_dict(ws)
        wb.close()
        with open(jsonfilename, "r", encoding='UTF-8') as f:
            jdata = json.load(f)
        lanlist = jdata['data']
        cnset = translateapi.create_cn_set(lanlist)
        newdict = {}
        for k in cnset.keys():
            cn = translateapi.escape_string(k)
            lan = dict1[cn] if cn in dict1 else None
            newdict[cn] =  lan        
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = u"中文"
        ws.cell(1, 2).value = u"译文"
        row = 2
        for k, v in newdict.items():
            ws.cell(row, 1).value = k
            ws.cell(row, 2).value = v
            row += 1
        (_, file_name) = os.path.split(excel1)
        (short_name, _) = os.path.splitext(file_name)
        save_dir = Dialog.asksaveasfilename(master = self.master, filetypes=[("Excel File", "*.xlsx")], title = "请选择保存Excel文件的路径", initialdir = config.excel_root(), initialfile = short_name)
        if save_dir == "":
            return
        (save_dir, file_name) = os.path.split(save_dir)
        (short_name, _) = os.path.splitext(file_name)
        save_full_name = os.path.join(save_dir, short_name + ".xlsx")
        wb.save(save_full_name)
        wb.close()

class DiffLan(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master=master)
        title1_2 = tk.Label(master=self, text = "新Excel:")
        new_excel = ui.FileOpenUI(master = self, file_types = [("Excel File", "*.xlsx")], base_root = lambda : config.excel_root(), title = u"翻译文件")
        # file1_2 = tk.StringVar(value = "")
        # input1_2 = tk.Entry(master=self, textvariable=file1_2)
        # btn1_2 = tk.Button(master=self, text="打开", command=lambda : viewtool.set_json_dir(self.winfo_toplevel(), file1_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=config.excel_root()))
        title2_2 = tk.Label(master=self, text = "旧Excel:")
        file2_2 = tk.StringVar(value = "")
        input2_2 = tk.Entry(master=self, textvariable=file2_2)
        btn2_2 = tk.Button(master=self, text="打开", command=lambda : viewtool.set_json_dir(self.winfo_toplevel(), file2_2, filetypes=[("Excel File", "*.xlsx")], title="翻译文件", initialdir=config.excel_root()))
        pickbtn = tk.Button(master=self, text="提取", command=lambda : self.diff_language(new_excel.get_value(), file2_2))
        row = 0    
        title1_2.grid(row = row, column = 0, columnspan = 6, sticky = tk.W)
        row += 1
        new_excel.grid(row = row, column = 0, columnspan = 6, sticky = tk.EW)
        # btn1_2.grid(row = row, column = 6, sticky = tk.EW, padx = 5)
        row += 1
        title2_2.grid(row = row, column = 0, columnspan = 6, sticky = tk.W)
        row += 1
        input2_2.grid(row = row, column = 0, columnspan = 6)
        btn2_2.grid(row = row, column = 6, sticky = tk.EW, padx = 5)
        row += 1
        pickbtn.grid(row = row, column = 0, sticky = tk.W)


    def diff_language(self, file1, file2):
        excel1 = file1.get()
        excel2 = file2.get()
        wb = openpyxl.load_workbook(excel1, read_only=True)
        ws = wb[wb.sheetnames[0]]
        dict1 = excel_to_dict(ws)
        wb.close()
        wb = openpyxl.load_workbook(excel2, read_only=True)
        ws = wb[wb.sheetnames[0]]
        dict2 = excel_to_dict(ws)
        wb.close()
        newdict = {}
        for k, v in dict1.items():
            if not k in dict2 or dict2[k] != v:
                newdict[k] = v 
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = u"中文"
        ws.cell(1, 2).value = u"译文"
        row = 2
        for k, v in newdict.items():
            ws.cell(row, 1).value = k
            ws.cell(row, 2).value = v
            row += 1
        (_, file_name) = os.path.split(excel1)
        (short_name, _) = os.path.splitext(file_name)
        save_dir = Dialog.asksaveasfilename(master = self.master, filetypes=[("Excel File", "*.xlsx")], title = "请选择保存Excel文件的路径", initialdir = config.excel_root(), initialfile = short_name)
        if save_dir == "":
            return
        (save_dir, file_name) = os.path.split(save_dir)
        (short_name, _) = os.path.splitext(file_name)
        save_full_name = os.path.join(save_dir, short_name + ".xlsx")
        wb.save(save_full_name)
        wb.close()

class TestView(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master=master, width = 300)
        ui.FileOpenUI(self).pack(fill = tk.X, expand = True)

class SkipView(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master=master)
        title = tk.Label(master=self, text = u"中文JSON:")
        jfile = ui.FileOpenUI(master = self, file_types = [("Json File", "*.json")], base_root = lambda:config.json_root(), title = u"中文JSON")
        title3_1 = tk.Label(master=self, text = u"忽略的表：")
        input3_1 = tk.Text(master=self, width = 80, height = 10)
        pickbtn = tk.Button(master=self, text="提取", command=lambda : self.skip_language(jfile.get_value(), input3_1))
        row = 0    
        title.grid(row = row, column = 0, columnspan = 6, sticky = tk.W)
        row += 1
        jfile.grid(row = row, column = 0, columnspan = 6, sticky = tk.EW)
        # btn1_2.grid(row = row, column = 6, sticky = tk.EW, padx = 5)
        
        row += 1
        title3_1.grid(row = row, column = 0, sticky = tk.W)
        row += 1
        input3_1.grid(row = row, column = 0, columnspan = 6, sticky = tk.NSEW)
        row += 1
        pickbtn.grid(row = row, column = 0, sticky = tk.W)


    def skip_language(self, jsonfilename, input_ui):
        root = self.winfo_toplevel()
        dbnamestring = input_ui.get("0.0", tk.END).strip()
        if jsonfilename == "" or not os.path.isfile(jsonfilename):
            MsgBox.showerror(u"错误", u"请选择一个json文件")
            return
        elif dbnamestring == "":
            MsgBox.showerror(u"错误", u"请输入要忽略的表名，用','隔开")
            return
        else:
            dbnames = dbnamestring.split(",")
            with open(jsonfilename, "r", encoding='UTF-8') as f:
                jdata = json.load(f)
                lanlist = jdata['data']
                def passed(item):
                    return item != None and item['table_name'] not in dbnames
                slimlist = [item for item in lanlist if passed(item)]
                jdata['data'] = slimlist
                (_, file_name) = os.path.split(jsonfilename)
                # (short_name, _) = os.path.splitext(file_name)
                save_dir = Dialog.asksaveasfilename(master = root, filetypes=[("Json File", "*.json")], title = "请选择保存json文件的路径", initialdir = config.json_root(), initialfile = file_name)
                if save_dir == "":
                    sytlog.log(u'取消\n')
                    return
                (save_dir, file_name) = os.path.split(save_dir)
                (short_name, _) = os.path.splitext(file_name)
                save_full_name = os.path.join(save_dir, short_name + ".json")
                with open(save_full_name, "w", encoding='UTF-8') as f:
                    json.dump(jdata, f, separators=(',',':'), ensure_ascii=False)
                    sytlog.log(u"成功.\n")

# if __name__ == "__main__":
#     DiffLan(TkinterDnD.Tk(), u"抽取未翻译的原文")