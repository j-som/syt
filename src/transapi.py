# 纯粹的翻译，不带那些特殊转换
import tkinter.filedialog as Dialog
import tkinter.messagebox as MsgBox
import openpyxl
import json
import re
import os.path as path
import tkinter as tk
json_root = "E:/st_kf/doc/language"
excel_root = "E:/sy_translation"
is_cn = re.compile("[\u4e00-\u9fbb]+")

def find_in_ws(lang, dictionary):
    if lang in dictionary:
        return dictionary[lang]
    return None

def collect_in_object(obj, cns):
    if type(obj) == dict:
        for _, item in obj.items():
            if type(item) == str:
                if is_cn.search(item):
                    cns[item] = cns[item] + 1 if item in cns else 1
            elif type(item) == dict or type(item) == list:
                collect_in_object(item, cns)
    elif type(obj) == list:
        for item in obj:
            if type(item) == str:
                if is_cn.search(item):
                    cns[item] = cns[item] + 1 if item in cns else 1
            elif type(item) == dict or type(item) == list:
                collect_in_object(item, cns)

def translate_in_obj(obj, dictionary):
    if type(obj) == dict:
        for key, item in obj.items():
            if type(item) == str:
                dest = find_in_ws(item, dictionary)
                if dest != None:
                    obj[key] = dest
            elif type(item) == dict or type(item) == list:
                translate_in_obj(item, dictionary)
    elif type(obj) == list:
        for i in range(0, len(obj)):
            item = obj[i]
            if type(item) == str:
                dest = find_in_ws(item, dictionary)
                if dest != None:
                    obj[i] = dest
            elif type(item) == dict or type(item) == list:
                translate_in_obj(item, dictionary)

def excel_to_dict(ws):
    dictionary = {}
    for i in ws.iter_rows():
        src = i[0].value
        if src != None:
            # print(src)
            dest = i[1].value
            dictionary[src] = dest
    return dictionary

def create_lan_excel_from_map(cnset, escape = True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = u"中文"
    ws.cell(1, 2).value = u"译文"
    row = 2
    for lang in cnset:
        ws.cell(row, 1).value = lang
        row += 1
    return wb

def create_cn_set(lanlist):
    cnset = {}
    collect_in_object(lanlist, cnset)
    return cnset

def translate_with_dict(lanlist, dictionary):
    translate_in_obj(lanlist, dictionary)