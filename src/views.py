import tkinter
import viewtool
import tkinter.messagebox as MsgBox
import tkinter.filedialog as Dialog
import json
import translateapi
import os
import openpyxl
import config
import sytlog
import tools 
import TkinterDnD2.TkinterDnD as TkinterDnD
def alert(func):
    def decorator(*args, **kw):
        try:
            func(*args, **kw)
        except Exception as e:
            sytlog.log(str(e))
            MsgBox.showerror(title = u"错误", message=str(e))
    return decorator


class PickView(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, master = parent)
        title1_1 = tkinter.Label(master=self, text = u"提取json源文件:")
        file1_1 = tkinter.StringVar(value = "")
        input1_1 = tkinter.Entry(master=self, textvariable=file1_1, width = 80)
        viewtool.bind_drop(input1_1)
        command = lambda : viewtool.set_json_dir(self.winfo_toplevel(), file1_1, filetypes=[("Json File", "*.json")], title=u"文件源json", initialdir=config.json_root())
        btn1_1 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        title1_2 = tkinter.Label(master=self, text = u"排除译文（可选，设置该路径将排除掉该Excel中已有的中文）:")
        file1_2 = tkinter.StringVar(value = "")
        input1_2 = tkinter.Entry(master=self, textvariable=file1_2, width = 80)
        viewtool.bind_drop(input1_2)
        command = lambda : viewtool.set_json_dir(self.winfo_toplevel(), file1_2, filetypes=[("Excel File", "*.xlsx")], title=u"翻译文件", initialdir=config.excel_root())
        btn1_2 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        pickbtn = tkinter.Button(master=self, text=u"提取", command=self.pick_cn)
        row = 0    
        title1_1.grid(row = row, column = 0, sticky = tkinter.W)
        row += 1
        input1_1.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn1_1.grid(row = row, column = 10, sticky = tkinter.EW, padx = 5)
        row += 1
        title1_2.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.W)
        row += 1
        input1_2.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn1_2.grid(row = row, column = 10, sticky = tkinter.EW, padx = 5)
        row += 1
        pickbtn.grid(row = row, column = 2, sticky = tkinter.EW)
        self._jsonfile = file1_1
        self._excelfile = file1_2
    @alert
    def pick_cn(self):
        root = self.winfo_toplevel()
        jsonfilename = self._jsonfile.get()
        if jsonfilename == "" or not os.path.isfile(jsonfilename):
            sytlog.log(u"没有选择json文件\n")
            MsgBox.showerror(u"错误", u"请选择一个json文件")
            return
        else:
            sytlog.log(u"开始提取\n")
            with open(jsonfilename, "r", encoding='UTF-8') as f:
                jdata = json.load(f)
            lanlist = jdata['data']
            cnset = translateapi.create_cn_set(lanlist)
            
            if len(cnset) > 0:
                excelfilename = self._excelfile.get()
                if excelfilename != "" and os.path.isfile(excelfilename):
                    wb1 = openpyxl.load_workbook(excelfilename, read_only = True)
                    dict1 = translateapi.excel_to_dict(wb1[wb1.sheetnames[0]])
                    wb1.close()
                    set1 = set()
                    for k in dict1:
                        set1.add(translateapi.invert_escape_string(k))
                    cnset = set(cnset) - set1
                    if len(cnset) == 0:
                        sytlog.log(u"没有可提取的中文\n")
                        return

                wb = translateapi.create_lan_excel_from_map(cnset)
                (_, file_name) = os.path.split(jsonfilename)
                (short_name, _) = os.path.splitext(file_name)
                save_dir = Dialog.asksaveasfilename(master = root, filetypes=[("Excel File", "*.xlsx")], title = "请选择保存Excel文件的路径", initialdir = config.excel_root(), initialfile = short_name + u".xlsx")
                if save_dir == "":
                    sytlog.log(u'取消\n')
                    return
                (save_dir, file_name) = os.path.split(save_dir)
                (short_name, _) = os.path.splitext(file_name)
                save_full_name = os.path.join(save_dir, short_name + ".xlsx")
                wb.save(save_full_name)
                wb.close()
                sytlog.log(u"提取完成\n")
            else:
                sytlog.log(u"没有可提取的中文\n")

class TransView(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, master = parent)
        root = self.winfo_toplevel()
        title2_1 = tkinter.Label(master=self, text = u"待翻译json文件:")
        file2_1 = tkinter.StringVar(value = "")
        self._jsonfile = file2_1
        input2_1 = tkinter.Entry(master=self, textvariable=file2_1)
        viewtool.bind_drop(input2_1)
        command = lambda : viewtool.set_json_dir(root, file2_1, filetypes=[("Json File", "*.json")], title=u"待翻译json", initialdir=config.json_root())
        btn2_1 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        row = 0
        title2_1.grid(row = row, column = 0, sticky = tkinter.W)
        row += 1
        input2_1.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn2_1.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)

        title2_2 = tkinter.Label(master=self, text = u"新译文:")
        file2_2 = tkinter.StringVar(value = "")
        self._excelfilenew = file2_2
        input2_2 = tkinter.Entry(master=self, textvariable=file2_2, width = 80)
        viewtool.bind_drop(input2_2)
        command = lambda : viewtool.set_json_dir(root, file2_2, filetypes=[("Excel File", "*.xlsx")], title=u"新译文", initialdir=config.excel_root())
        btn2_2 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        row += 1
        title2_2.grid(row = row, column = 0, sticky = tkinter.W)
        row += 1
        input2_2.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn2_2.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)

        title3_2 = tkinter.Label(master=self, text = u"旧译文:（可选，假如存在，则会把json中的旧译文替换成新译文）")
        file3_2 = tkinter.StringVar(value = "")
        self._excelfileold = file3_2
        input3_2 = tkinter.Entry(master=self, textvariable=file3_2, width = 80)
        viewtool.bind_drop(input3_2)
        command = lambda : viewtool.set_json_dir(root, file3_2, filetypes=[("Excel File", "*.xlsx")], title=u"旧译文", initialdir=config.excel_root())
        btn3_2 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        row += 1
        title3_2.grid(row = row, column = 0, sticky = tkinter.W, columnspan = 6)
        row += 1
        input3_2.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn3_2.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)

        transbtn = tkinter.Button(master=self, text=u"翻译", command=self.translate)
        row += 1
        transbtn.grid(row = row, column = 2, sticky = tkinter.EW)
    @alert
    def translate(self):
        root = self.winfo_toplevel()
        jsonfilename = self._jsonfile.get()
        if jsonfilename == "" or not os.path.isfile(jsonfilename):
            MsgBox.showerror(u"错误", u"请选择一个json文件")
            return
        else:
            excelfilename = self._excelfilenew.get()
            if excelfilename == "" or not os.path.isfile(excelfilename):
                MsgBox.showerror(u"错误", u"请选择一个Excel文件")
                return 
            sytlog.log(u"开始翻译\n")
            with open(jsonfilename, "r", encoding='UTF-8') as f:
                jdata = json.load(f)

            lanlist = jdata['data']
            ref_workbook=openpyxl.load_workbook(excelfilename, read_only = True)
            ws = ref_workbook[ref_workbook.sheetnames[0]]# index为0为第一张表 
            dictionary = translateapi.excel_to_dict(ws)
            ref_workbook.close()
            translateapi.translate_with_dict(lanlist, dictionary)
            excelfilename2 = self._excelfileold.get()
            if excelfilename2 != "" and os.path.isfile(excelfilename2):
                ref_workbook = openpyxl.load_workbook(excelfilename2, read_only = True)
                ws = ref_workbook[ref_workbook.sheetnames[0]]
                old_dict = translateapi.excel_to_dict(ws)
                update_dict = translateapi.make_update_dict(old_dict, dictionary)
                if len(update_dict) > 0:
                    # print(update_dict)
                    translateapi.translate_with_dict(lanlist, update_dict)
                else:
                    sytlog.log(u"无更新的译文\n")

            (_, file_name) = os.path.split(jsonfilename)
            # (short_name, _) = os.path.splitext(file_name)
            save_dir = Dialog.asksaveasfilename(master = root, filetypes=[("Json File", "*.json")], title = u"请选择保存json文件的路径", initialdir = config.json_root(), initialfile = file_name)
            if save_dir == "":
                sytlog.log(u'取消\n')
                return
            (save_dir, file_name) = os.path.split(save_dir)
            (short_name, _) = os.path.splitext(file_name)
            save_full_name = os.path.join(save_dir, short_name + ".json")
            with open(save_full_name, "w", encoding='UTF-8') as f:
                json.dump(jdata, f, separators=(',',':'), ensure_ascii=False)
                sytlog.log(u"翻译完成.\n")

class UpdateView(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, master = parent)
        root = self.winfo_toplevel()
        title2_1 = tkinter.Label(master=self, text = u"待翻译json文件:")
        file2_1 = tkinter.StringVar(value = "")
        self._jsonfile = file2_1
        input2_1 = tkinter.Entry(master=self, textvariable=file2_1)
        viewtool.bind_drop(input2_1)
        command = lambda : viewtool.set_json_dir(root, file2_1, filetypes=[("Json File", "*.json")], title=u"待翻译json", initialdir=config.json_root())
        btn2_1 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        row = 0
        title2_1.grid(row = row, column = 0, sticky = tkinter.W)
        row += 1
        input2_1.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn2_1.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)

        title2_2 = tkinter.Label(master=self, text = u"新译文:")
        file2_2 = tkinter.StringVar(value = "")
        self._excelfilenew = file2_2
        input2_2 = tkinter.Entry(master=self, textvariable=file2_2, width = 80)
        viewtool.bind_drop(input2_2)
        command = lambda : viewtool.set_json_dir(root, file2_2, filetypes=[("Excel File", "*.xlsx")], title=u"新译文", initialdir=config.excel_root())
        btn2_2 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        row += 1
        title2_2.grid(row = row, column = 0, sticky = tkinter.W)
        row += 1
        input2_2.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn2_2.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)

        title3_2 = tkinter.Label(master=self, text = u"旧译文:")
        file3_2 = tkinter.StringVar(value = "")
        self._excelfileold = file3_2
        input3_2 = tkinter.Entry(master=self, textvariable=file3_2, width = 80)
        viewtool.bind_drop(input3_2)
        command = lambda : viewtool.set_json_dir(root, file3_2, filetypes=[("Excel File", "*.xlsx")], title=u"旧译文", initialdir=config.excel_root())
        btn3_2 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        row += 1
        title3_2.grid(row = row, column = 0, sticky = tkinter.W, columnspan = 6)
        row += 1
        input3_2.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn3_2.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)

        transbtn = tkinter.Button(master=self, text=u"更新", command=self.update)
        row += 1
        transbtn.grid(row = row, column = 2, sticky = tkinter.EW)

    # @alert
    def update(self):
        root = self.winfo_toplevel()
        jsonfilename = self._jsonfile.get()
        if jsonfilename == "" or not os.path.isfile(jsonfilename):
            MsgBox.showerror(u"错误", u"请选择一个json文件")
            return
        else:
            excelfilename = self._excelfilenew.get()
            excelfilename2 = self._excelfileold.get()
            if excelfilename == "" or not os.path.isfile(excelfilename) or excelfilename2 == "" or not os.path.isfile(excelfilename2):
                MsgBox.showerror(u"错误", u"请选择好Excel文件")
                return 
            sytlog.log(u"开始更新\n")
            with open(jsonfilename, "r", encoding='UTF-8') as f:
                jdata = json.load(f)

            lanlist = jdata['data']
            ref_workbook=openpyxl.load_workbook(excelfilename, read_only = True)
            ws = ref_workbook[ref_workbook.sheetnames[0]]# index为0为第一张表 
            dictionary = translateapi.excel_to_dict(ws)
            ref_workbook.close()
            ref_workbook = openpyxl.load_workbook(excelfilename2, read_only = True)
            ws = ref_workbook[ref_workbook.sheetnames[0]]
            old_dict = translateapi.excel_to_dict(ws)
            ref_workbook.close()
            update_dict = translateapi.make_update_dict(old_dict, dictionary)
            if len(update_dict) > 0:
                newlist = translateapi.update_with_dict(lanlist, update_dict)
                if len(newlist) > 0:
                    jdata['data'] = newlist
                    (_, file_name) = os.path.split(jsonfilename)
                    # (short_name, _) = os.path.splitext(file_name)
                    save_dir = Dialog.asksaveasfilename(master = root, filetypes=[("Json File", "*.json")], title = "请选择保存json文件的路径", initialdir = config.json_root(), initialfile = file_name)
                    if save_dir == "":
                        sytlog.log(u'取消\n')
                        return
                    (save_dir, file_name) = os.path.split(save_dir)
                    (short_name, _) = os.path.splitext(file_name)
                    save_full_name = os.path.join(save_dir, short_name + ".json")
                    with open(save_full_name, "w", encoding='UTF-8') as f:
                        json.dump(jdata, f, separators=(',',':'), ensure_ascii=False)
                        sytlog.log(u"更新完成.\n")
                else:
                    sytlog.log(u"找不到更新的译文\n" + str(update_dict.keys()) + '\n')
            else:
                sytlog.log(u"无更新的译文\n")

class TwView(tkinter.Frame):
    def __init__(self, parent):
        tkinter.Frame.__init__(self, master = parent)
        root = self.winfo_toplevel()
        title2_1 = tkinter.Label(master=self, text = u"待翻译json文件:")
        file2_1 = tkinter.StringVar(value = "")
        self._jsonfile = file2_1
        input2_1 = tkinter.Entry(master=self, textvariable=file2_1, width = 80)
        viewtool.bind_drop(input2_1)
        command = lambda : viewtool.set_json_dir(root, file2_1, filetypes=[("Json File", "*.json")], title=u"待翻译json", initialdir=config.json_root())
        btn2_1 = tkinter.Button(master=self, text=u"打开", command=command, width = 10)
        row = 0
        title2_1.grid(row = row, column = 0, sticky = tkinter.W)
        row += 1
        input2_1.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.EW)
        btn2_1.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)
        transbtn = tkinter.Button(master=self, text=u"机译", command=self.translate)
        row += 1
        transbtn.grid(row = row, column = 2, sticky = tkinter.EW)
        title3_1 = tkinter.Label(master=self, text = u"原文：")
        # self._srcword = file3_1
        input3_1 = tkinter.Text(master=self, width = 80, height = 10)
        command = lambda : sytlog.log(translateapi.twconvertor.convert(input3_1.get("0.0", tkinter.END)))
        btn3_1 = tkinter.Button(master=self, text=u"翻译", command=command, width = 10)
        row += 1
        title3_1.grid(row = row, column = 0, sticky = tkinter.W)
        row += 1
        input3_1.grid(row = row, column = 0, columnspan = 6, sticky = tkinter.NSEW)
        btn3_1.grid(row = row, column = 6, sticky = tkinter.EW, padx = 5)

    @alert
    def translate(self):
        root = self.winfo_toplevel()
        jsonfilename = self._jsonfile.get()
        if jsonfilename == "" or not os.path.isfile(jsonfilename):
            MsgBox.showerror(u"错误", u"请选择一个json文件")
            return
        else:
            with open(jsonfilename, "r", encoding='UTF-8') as f:
                jdata = json.load(f)
            lanlist = jdata['data']
            translateapi.translate2tw(lanlist)
            (_, file_name) = os.path.split(jsonfilename)
            # (short_name, _) = os.path.splitext(file_name)
            save_dir = Dialog.asksaveasfilename(master = root, filetypes=[("Json File", "*.json")], title = u"请选择保存json文件的路径", initialdir = config.json_root(), initialfile = file_name)
            if save_dir == "":
                sytlog.log(u'取消\n')
                return
            (save_dir, file_name) = os.path.split(save_dir)
            (short_name, _) = os.path.splitext(file_name)
            save_full_name = os.path.join(save_dir, short_name + ".json")
            with open(save_full_name, "w", encoding='UTF-8') as f:
                json.dump(jdata, f, separators=(',',':'), ensure_ascii=False)
                sytlog.log(u"翻译完成.\n")


class _btn_class(tkinter.Button):
    def __init__(self, master, text, view_class, view_container):
        tkinter.Button.__init__(self, master=master, text = text)
        self.config(bg = '#dddddd')
        self._viewcontainer = view_container
        self._viewcls = view_class
        self._view = None
    
    def select(self):
        self.config(bg = '#ffffff')
        if self._view == None:
            self._view = self._viewcls(self._viewcontainer)
        self._view.pack()

    def unselect(self):
        self.config(bg = '#dddddd')
        self._view.pack_forget()

    def __delete__(self, instance):
        self._view.destory()
        self.__delete__(instance)
        
class TabsView(tkinter.Frame):

    def __init__(self, master, column = 255):
        tkinter.Frame.__init__(self, master=master)
        self._container = tkinter.PanedWindow(master=self)
        self._tabbtns = []
        self.curbtn = None
        self._btncontainer = tkinter.Frame(master=self)
        self._btncontainer.pack(anchor = tkinter.W)
        self._column = column
        self._container.pack(side = tkinter.LEFT)

    def add(self, label, viewclass):
        btn = _btn_class(self._btncontainer, label, viewclass, self._container)
        self._tabbtns.append(btn)
        index = len(self._tabbtns) - 1
        btn.grid(row = index // self._column, column = index % self._column, ipadx = 10, sticky = tkinter.NSEW)
        btn.bind(sequence='<Button-1>', func=self._btn_select)
        if not self.curbtn:
            btn.select()
            self.curbtn = btn 

    def _btn_select(self, event):
        btn = event.widget
        if btn == self.curbtn: 
            return
        if self.curbtn != None:
            self.curbtn.unselect()

        btn.select()
        self.curbtn = btn

    
class LogPanel(tkinter.Frame):
    def __init__(self, master):
        tkinter.Frame.__init__(self, master = master)
        _container = tkinter.Frame(master=self)
        _scrollbar = tkinter.Scrollbar(master=_container)
        _scrollbar.pack(side = tkinter.RIGHT, fill=tkinter.Y)
        self._textpanel = tkinter.Text(master=_container, yscrollcommand=_scrollbar.set)
        self._clearbtn = tkinter.Button(master=self, text=u'Clear', command = self._clear_text)
        self._clearbtn.pack(anchor = tkinter.E)
        self._textpanel.pack(expand = True, fill = tkinter.BOTH)
        _container.pack(expand = True, fill = tkinter.BOTH)
        _scrollbar.config(command = self._textpanel.yview)

    def _clear_text(self):
        self._textpanel.delete(0.0, tkinter.END)

    def print(self, text):
        self._textpanel.insert(tkinter.END, text)

class CompErl(tkinter.Frame):
    def __init__(self, master):
        tkinter.Frame.__init__(self, master=master)

class ToolsView(tkinter.Frame):
    def __init__(self, master):
        tkinter.Frame.__init__(self, master = master)
        tabframe = TabsView(master=self, column = 2)
        tabframe.pack(anchor = tkinter.W)
        tabs = {
            u'抽取未翻译的原文':tools.DiffLan,
            u'删除无用的翻译':tools.SlimLan,
            u'测试':tools.TestView
            }
        for tabname in tabs:
            tabframe.add(tabname, tabs[tabname])

