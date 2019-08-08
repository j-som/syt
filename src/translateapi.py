import tkinter.filedialog as Dialog
import tkinter.messagebox as MsgBox
import openpyxl
import json
import re
import os.path as path
import tkinter as tk
import sytlog 
import langconv
json_root = "E:/st_kf/doc/language"
excel_root = "E:/sy_translation"
is_cn = re.compile("[\u4e00-\u9fa5]+")

# 以前json提取出来的文本带着转义字符
# 而json文件解析后的文本不带转义
# 所以需要将解析后的lang加上转义字符后，再和提取的文本对比
def escape_string(lang):
    return json.dumps(lang, ensure_ascii=False)[1:-1]

def invert_escape_string(lang):
    if type(lang) == str:
        try:
            return json.loads(r'"' + lang + r'"')
        except:
            sytlog.log(u"请检查：[" + str(lang) + u']的格式\n')
            return lang
    return lang 


def find_in_ws(lang, dictionary):
    lang = escape_string(lang)
    if lang in dictionary:
        v = dictionary[lang]
        if type(v) == str:
            return v.strip()
        else: 
            return v
    return None

def collect_in_object(obj, cns):
    if type(obj) == dict:
        for _, item in obj.items():
            if type(item) == str:
                if is_cn.search(item):
                    cns[item] = cns[item] + 1 if item in cns else 1
            elif type(item) == dict or type(item) == list:
                collect_in_object(item, cns)
    elif type(obj) == list:
        for item in obj:
            if type(item) == str:
                if is_cn.search(item):
                    cns[item] = cns[item] + 1 if item in cns else 1
            elif type(item) == dict or type(item) == list:
                collect_in_object(item, cns)

def translate_in_obj(obj, dictionary):
    if type(obj) == dict:
        for key, item in obj.items():
            if type(item) == str:
                dest = find_in_ws(item, dictionary)
                if dest != None:
                    obj[key] = invert_escape_string(dest)
            elif type(item) == dict or type(item) == list:
                translate_in_obj(item, dictionary)
    elif type(obj) == list:
        for i in range(0, len(obj)):
            item = obj[i]
            if type(item) == str:
                dest = find_in_ws(item, dictionary)
                if dest != None:
                    obj[i] = invert_escape_string(dest)
            elif type(item) == dict or type(item) == list:
                translate_in_obj(item, dictionary)

def update_in_obj(obj, dictionary, flag):
    if type(obj) == dict:
        for key, item in obj.items():
            if type(item) == str:
                dest = find_in_ws(item, dictionary)
                if dest != None:
                    obj[key] = invert_escape_string(dest)
                    flag += 1
            elif type(item) == dict or type(item) == list:
                flag = update_in_obj(item, dictionary, flag)
    elif type(obj) == list:
        for i in range(0, len(obj)):
            item = obj[i]
            if type(item) == str:
                dest = find_in_ws(item, dictionary)
                if dest != None:
                    obj[i] = invert_escape_string(dest)
                    flag += 1
            elif type(item) == dict or type(item) == list:
                flag = update_in_obj(item, dictionary, flag)
    return flag

def fix_error(dest):
    dest = dest.replace(r'\"', r'e-s-c-a-p-e-dou-ble--q-uote')
    # 不带转义\的双引号在配置的生成那里会引起语法错误
    dest = dest.replace(r'"', r'＂')
    dest = dest.replace(r'e-s-c-a-p-e-dou-ble--q-uote', r'\"')
    # 换行连接符后面不能有空白字符
    dest = re.sub(r'(\\)[\s]+', lambda m:m.group(1), dest)
    return dest 

def excel_to_dict(ws):
    dictionary = {}
    for i in ws.iter_rows():
        src = i[0].value
        if src != None:
            # 后台提取出来的/带有\，被json.load进入后会被删除
            # 为了保持一致，尝试删除原文中的\
            src = src.replace(r'\/', r'/')
            # print(src)
            dest = i[1].value
            if type(dest) == str:
                dest = fix_error(dest)
            dictionary[src] = dest
    return dictionary

def create_lan_excel_from_map(cnset, escape = True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = u"中文"
    ws.cell(1, 2).value = u"译文"
    row = 2
    for lang in cnset:
        ws.cell(row, 1).value = escape_string(lang) if escape else lang
        row += 1
    return wb

def create_cn_set(lanlist):
    cnset = {}
    for item in lanlist:
        obj = item['lang']
        if type(obj) == str:
            if is_cn.search(obj):
                cnset[obj] = cnset[obj] + 1 if obj in cnset else 1
        else:
            collect_in_object(obj, cnset)
    return cnset

def translate_with_dict(lanlist, dictionary):
    for item in lanlist:
        obj = item['lang']
        if type(obj) == str:
            dest = find_in_ws(obj, dictionary)
            if dest != None and dest != "":
                dest = invert_escape_string(dest)
                item['lang'] = dest
            elif is_cn.search(obj):
                sytlog.log(u"[" + obj + u"]未找到译文\n")
        else: 
            translate_in_obj(obj, dictionary)

def update_with_dict(lanlist, dictionary):
    updatelist = []
    for item in lanlist:
        obj = item['lang']
        if type(obj) == str:
            dest = find_in_ws(obj, dictionary)
            if dest != None:
                dest = invert_escape_string(dest)
                item['lang'] = dest
                updatelist.append(item)
        else: 
            change = update_in_obj(obj, dictionary, 0)
            if change > 0:
                updatelist.append(item)
    return updatelist

def make_update_dict(old_dict, new_dict):
    update_dict = {}
    sytlog.log("----------------------------------------\n")
    for cn, new in new_dict.items():
        if cn in old_dict:
            old = old_dict[cn]
            if old != new:
                old = invert_escape_string(old)
                old = escape_string(old)
                update_dict[old] = new
                sytlog.log(cn + "\n" + old + "\n" + new + "\n\n")
        else:
            cn = invert_escape_string(cn)
            cn = escape_string(cn)
            update_dict[cn] = new
            sytlog.log(cn + "\n" + new + "\n\n")
    sytlog.log("----------------------------------------\n") 

    return update_dict


from zh_wiki import zh2TW, zh2Hant
zh2tw = {}
zh2tw.update(zh2Hant)
zh2tw.update(zh2TW)
langconv.registery('zh-tw', zh2tw)
del zh2TW, zh2Hant, zh2tw
twconvertor = langconv.Converter('zh-tw')

def simple2tradition(line):
    #将简体转换成繁体
    line = langconv.Converter('zh-tw').convert(line)
    # line = line.encode('utf-8')
    return line
 
def tradition2simple(line):
    # 将繁体转换成简体
    line = langconv.Converter('zh-hans').convert(line)
    # line = line.encode('utf-8')
    return line

def translate2tw(lanlist):
    global twconvertor
    for item in lanlist:
        obj = item['lang']
        if type(obj) == str:
            item['lang'] = twconvertor.convert(obj)
        else: 
            translate2tw_in_obj(obj)
    
def translate2tw_in_obj(obj):
    global twconvertor
    if type(obj) == dict:
        for key, item in obj.items():
            if type(item) == str:
                obj[key] = twconvertor.convert(item)
            elif type(item) == dict or type(item) == list:
                translate2tw_in_obj(item)
    elif type(obj) == list:
        for i in range(0, len(obj)):
            item = obj[i]
            if type(item) == str:
                obj[i] = twconvertor.convert(item)
            elif type(item) == dict or type(item) == list:
                translate2tw_in_obj(item)